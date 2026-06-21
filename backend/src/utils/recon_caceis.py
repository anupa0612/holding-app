from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _norm(s: str) -> str:
    return "".join(ch for ch in s.strip().lower() if ch.isalnum())


def _to_number(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return 0.0
    s = s.replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except Exception:
        return 0.0


def _decode_csv_bytes(raw: bytes) -> str:
    # Honour an explicit BOM first (handles real UTF-16/UTF-8 files).
    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        return raw.decode("utf-16")
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig")

    # Otherwise prefer single-byte encodings. UTF-16 is intentionally NOT in the
    # blind fallback list because it "succeeds" on most byte sequences and turns
    # CP1252/Latin-1 files (e.g. accented characters like "Æ") into garbage.
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _read_rows(path: str) -> list[dict[str, Any]]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        raw = Path(path).read_bytes()
        text = _decode_csv_bytes(raw)

        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except Exception:
            dialect = csv.excel

        f = io.StringIO(text)
        reader = csv.DictReader(f, dialect=dialect)
        return [dict(r) for r in reader]
    if ext in {".xlsx", ".xls"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = next(it, None) or []
        cols = [str(h).strip() if h is not None else "" for h in headers]
        rows: list[dict[str, Any]] = []
        for row in it:
            rows.append({cols[i]: (row[i] if i < len(row) else None) for i in range(len(cols))})
        return rows
    raise ValueError("Unsupported file type. Please upload CSV or Excel.")


def _pick(row: dict[str, Any], candidates: list[str]) -> Any:
    norm_map = {_norm(k): k for k in row.keys()}
    for c in candidates:
        k = norm_map.get(_norm(c))
        if k is not None:
            return row.get(k)
    return None


@dataclass(frozen=True)
class HoldingsConfig:
    """Column mapping for position holdings reconciliation (AT + broker files)."""

    broker_label: str
    cp_isin_columns: tuple[str, ...]
    cp_qty_columns: tuple[str, ...]
    at_isin_columns: tuple[str, ...] = ("ISIN", "ISIN Code", "ISIN CODE")
    at_qty_columns: tuple[str, ...] = ("Settled Quantity", "SettledQty", "Quantity")
    at_customer_no_columns: tuple[str, ...] = ("Customer No", "CustomerNo", "Customer Number")
    at_customer_name_columns: tuple[str, ...] = ("Customer Name", "CustomerName")
    at_date_columns: tuple[str, ...] = ("Date",)
    at_exchange_columns: tuple[str, ...] = ("Exchange",)
    # Decimal places used to decide matched vs break and to display the difference.
    # 5 keeps full precision (CACEIS/Clear Street); 2 means a line is matched when
    # the difference is zero at 2 decimals (i.e. sub-0.01 rounding noise is ignored).
    match_decimals: int = 5


CACEIS_CP = HoldingsConfig(
    broker_label="CACEIS",
    cp_isin_columns=("ISIN CODE", "ISIN", "ISIN Code"),
    cp_qty_columns=("BALANCE AT RECORD DATE", "BALANCE", "Balance at record date"),
    at_isin_columns=("ISIN Code", "ISIN", "ISIN CODE"),
)

CLEARSTREET_CP = HoldingsConfig(
    broker_label="Clear Street",
    cp_isin_columns=("ISIN", "ISIN CODE", "ISIN Code"),
    cp_qty_columns=("SD Quantity", "SD Qty", "SD QUANTITY"),
)

GTNA_HOLDINGS = HoldingsConfig(
    broker_label="GTNA",
    cp_isin_columns=("ISIN_CODE", "ISIN CODE", "ISIN Code", "ISIN"),
    cp_qty_columns=("SETTLED_QUANTITY", "Settled Quantity", "SETTLED QUANTITY"),
    at_isin_columns=("ISIN Code", "ISIN_CODE", "ISIN CODE", "ISIN", "ISINCode"),
    at_qty_columns=("Settled Quantity", "SETTLED_QUANTITY", "SettledQty", "Settled Qty", "Quantity"),
    at_customer_no_columns=("Customer No", "CustomerNo", "Customer Number", "Customer No.", "Cust No"),
    at_customer_name_columns=("Customer Name", "CustomerName", "Cust Name"),
    at_date_columns=("Date", "Settlement Date", "Settled Date"),
    match_decimals=2,
)

GTNME_HOLDINGS = HoldingsConfig(
    broker_label="GTNME",
    cp_isin_columns=("ISIN Code", "ISIN CODE", "ISIN", "ISIN_CODE"),
    cp_qty_columns=("Net Holdings", "Net holdings", "NET HOLDINGS"),
    at_isin_columns=("ISIN Code", "ISIN_CODE", "ISIN CODE", "ISIN", "ISINCode"),
    at_qty_columns=("Net Holdings", "Net holdings", "NET HOLDINGS", "Settled Quantity", "Quantity"),
    at_customer_no_columns=("Customer No", "CustomerNo", "Customer Number", "Customer No.", "Cust No"),
    at_customer_name_columns=("Customer Name", "CustomerName", "Cust Name"),
    at_date_columns=("Date", "Settlement Date", "Settled Date"),
    match_decimals=2,
)

# Backward-compatible alias for template key eu_settled_holdings
EU_SETTLED_CP = GTNA_HOLDINGS


@dataclass(frozen=True)
class CaceisResult:
    matched: list[dict[str, Any]]
    breaks: list[dict[str, Any]]
    only_our: list[dict[str, Any]]
    only_cp: list[dict[str, Any]]
    summary: dict[str, Any]


def _holdings_stats(our_path: str, cp_path: str, cfg: HoldingsConfig) -> dict[str, Any]:
    our_rows = _read_rows(our_path)
    cp_rows = _read_rows(cp_path)

    ROUND_DP = 5

    def r5(x: float) -> float:
        return float(f"{x:.{ROUND_DP}f}")

    our_total = 0.0
    for r in our_rows:
        our_total += _to_number(_pick(r, list(cfg.at_qty_columns)))

    cp_total = 0.0
    for r in cp_rows:
        cp_total += _to_number(_pick(r, list(cfg.cp_qty_columns)))

    diff_raw = our_total - cp_total
    diff_5 = r5(diff_raw)

    return {
        "ourLineCount": len(our_rows),
        "cpLineCount": len(cp_rows),
        "ourHoldingTotal": our_total,
        "cpHoldingTotal": cp_total,
        "breakValue": diff_5,
        "breakValueRaw": diff_raw,
    }


def _build_holdings(our_path: str, cp_path: str, cfg: HoldingsConfig) -> CaceisResult:
    """
    AT report (our): columns per HoldingsConfig (GTNA: Date, Customer No, Customer Name, ISIN Code, Settled Quantity)
    Multiple AT lines with the same ISIN + customer are summed; broker lines per ISIN are summed too.
    Match key: ISIN
    Net holding difference: our_settled_qty - broker_balance
    """
    our_rows = _read_rows(our_path)
    cp_rows = _read_rows(cp_path)

    broker_balance_key = f"Balance at record date ({cfg.broker_label} - ISIN total)"
    broker_line_count_key = f"{cfg.broker_label} line count"

    ROUND_DP = 5
    MATCH_DP = cfg.match_decimals

    def r5(x: float) -> float:
        return float(f"{x:.{MATCH_DP}f}")

    def diff_display(raw: float, rounded5: float) -> str | float:
        # Coarse tolerance (e.g. GTNA 2dp): show the rounded difference only.
        if MATCH_DP < 5:
            return float(f"{raw:.{MATCH_DP}f}")
        if rounded5 != 0.0:
            return f"{raw:.15f}".rstrip("0").rstrip(".")
        return rounded5

    our_by_customer: dict[tuple[str, str, str], dict[str, Any]] = {}
    for r in our_rows:
        isin = _pick(r, list(cfg.at_isin_columns))
        if isin is None:
            continue
        isin_s = str(isin).strip().upper()
        if not isin_s:
            continue

        cust_no = str(_pick(r, list(cfg.at_customer_no_columns)) or "").strip()
        cust_name = str(_pick(r, list(cfg.at_customer_name_columns)) or "").strip()
        cust_key = (isin_s, cust_no, cust_name)

        qty = _to_number(_pick(r, list(cfg.at_qty_columns)))
        cur = our_by_customer.get(cust_key)
        if not cur:
            our_by_customer[cust_key] = {
                "ISIN": isin_s,
                "Date": _pick(r, list(cfg.at_date_columns)),
                "Customer No": cust_no,
                "Customer Name": cust_name,
                "Exchange": _pick(r, list(cfg.at_exchange_columns)),
                "Settled Quantity (AT)": qty,
                "AT line count": 1,
            }
        else:
            cur["Settled Quantity (AT)"] = float(cur.get("Settled Quantity (AT)", 0.0)) + qty
            cur["AT line count"] = int(cur.get("AT line count", 0)) + 1

    our_total_by_isin: dict[str, float] = {}
    our_rows_by_isin: dict[str, list[dict[str, Any]]] = {}
    for (isin, _, _), row in our_by_customer.items():
        our_total_by_isin[isin] = our_total_by_isin.get(isin, 0.0) + float(row.get("Settled Quantity (AT)", 0.0))
        our_rows_by_isin.setdefault(isin, []).append(row)

    cp_agg: dict[str, dict[str, Any]] = {}
    for r in cp_rows:
        isin = _pick(r, list(cfg.cp_isin_columns))
        if isin is None:
            continue
        isin_s = str(isin).strip().upper()
        if not isin_s:
            continue

        bal = _to_number(_pick(r, list(cfg.cp_qty_columns)))
        cur = cp_agg.get(isin_s)
        if not cur:
            cp_agg[isin_s] = {
                "ISIN": isin_s,
                broker_balance_key: bal,
                broker_line_count_key: 1,
            }
        else:
            cur[broker_balance_key] = float(cur.get(broker_balance_key, 0.0)) + bal
            cur[broker_line_count_key] = int(cur.get(broker_line_count_key, 0)) + 1

    all_isins = sorted(set(our_total_by_isin.keys()) | set(cp_agg.keys()))

    matched: list[dict[str, Any]] = []
    breaks: list[dict[str, Any]] = []
    only_our: list[dict[str, Any]] = []
    only_cp: list[dict[str, Any]] = []

    for isin in all_isins:
        our_total = float(our_total_by_isin.get(isin, 0.0))
        c = cp_agg.get(isin)
        if isin in our_total_by_isin and c:
            broker_total = float(c.get(broker_balance_key, 0.0))
            # Skip ISINs with no holdings on either side (0 AT and 0 broker).
            if r5(our_total) == 0.0 and r5(broker_total) == 0.0:
                continue
            isin_diff_raw = our_total - broker_total
            isin_diff_5 = r5(isin_diff_raw)

            for o in our_rows_by_isin.get(isin, []):
                row = {
                    "ISIN": isin,
                    "Date": o.get("Date"),
                    "Customer No": o.get("Customer No"),
                    "Customer Name": o.get("Customer Name"),
                    "Exchange": o.get("Exchange"),
                    "Settled Quantity (AT - customer)": float(o.get("Settled Quantity (AT)", 0.0)),
                    "AT line count": o.get("AT line count", 1),
                    "AT total (ISIN)": our_total,
                    broker_balance_key: broker_total,
                    broker_line_count_key: c.get(broker_line_count_key, 1),
                    "Net holding difference (ISIN total)": isin_diff_5,
                    "Net holding difference (raw)": isin_diff_raw,
                    "Net holding difference (display)": diff_display(isin_diff_raw, isin_diff_5),
                }
                if isin_diff_5 == 0.0:
                    matched.append(row)
                else:
                    breaks.append(row)
        elif isin in our_total_by_isin and not c:
            if r5(our_total) == 0.0:
                continue
            for o in our_rows_by_isin.get(isin, []):
                only_our.append(
                    {
                        "ISIN": isin,
                        "Date": o.get("Date"),
                        "Customer No": o.get("Customer No"),
                        "Customer Name": o.get("Customer Name"),
                        "Exchange": o.get("Exchange"),
                        "Settled Quantity (AT - customer)": float(o.get("Settled Quantity (AT)", 0.0)),
                        "AT line count": o.get("AT line count", 1),
                        "AT total (ISIN)": our_total,
                    }
                )
        elif c and isin not in our_total_by_isin:
            if r5(float(c.get(broker_balance_key, 0.0))) == 0.0:
                continue
            only_cp.append(
                {
                    "ISIN": isin,
                    broker_balance_key: float(c.get(broker_balance_key, 0.0)),
                    broker_line_count_key: c.get(broker_line_count_key, 1),
                }
            )

    summary = {
        "matched": 0,
        "breaks": 0,
        "onlyOur": 0,
        "onlyCp": 0,
        "totalIsins": len(all_isins),
    }
    for isin in all_isins:
        in_our = isin in our_total_by_isin
        in_cp = isin in cp_agg
        our_t = float(our_total_by_isin.get(isin, 0.0))
        broker_t = float((cp_agg.get(isin) or {}).get(broker_balance_key, 0.0))
        if in_our and in_cp:
            # Skip ISINs with no holdings on either side.
            if r5(our_t) == 0.0 and r5(broker_t) == 0.0:
                continue
            if r5(our_t - broker_t) == 0.0:
                summary["matched"] += 1
            else:
                summary["breaks"] += 1
        elif in_our and not in_cp:
            if r5(our_t) == 0.0:
                continue
            summary["onlyOur"] += 1
        elif in_cp and not in_our:
            if r5(broker_t) == 0.0:
                continue
            summary["onlyCp"] += 1

    return CaceisResult(
        matched=matched,
        breaks=breaks,
        only_our=only_our,
        only_cp=only_cp,
        summary=summary,
    )


def caceis_stats(our_path: str, cp_path: str) -> dict[str, Any]:
    return _holdings_stats(our_path, cp_path, CACEIS_CP)


def clearstreet_stats(our_path: str, cp_path: str) -> dict[str, Any]:
    return _holdings_stats(our_path, cp_path, CLEARSTREET_CP)


def eu_settled_stats(our_path: str, cp_path: str) -> dict[str, Any]:
    return _holdings_stats(our_path, cp_path, GTNA_HOLDINGS)


def gtnme_stats(our_path: str, cp_path: str) -> dict[str, Any]:
    return _holdings_stats(our_path, cp_path, GTNME_HOLDINGS)


def build_caceis_holdings(our_path: str, cp_path: str) -> CaceisResult:
    return _build_holdings(our_path, cp_path, CACEIS_CP)


def build_clearstreet_holdings(our_path: str, cp_path: str) -> CaceisResult:
    return _build_holdings(our_path, cp_path, CLEARSTREET_CP)


def build_eu_settled_holdings(our_path: str, cp_path: str) -> CaceisResult:
    return _build_holdings(our_path, cp_path, GTNA_HOLDINGS)


def build_gtnme_holdings(our_path: str, cp_path: str) -> CaceisResult:
    return _build_holdings(our_path, cp_path, GTNME_HOLDINGS)
