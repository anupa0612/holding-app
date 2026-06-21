from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import csv
import io
from openpyxl import load_workbook
from flask import current_app


Side = Literal["our", "cp"]


@dataclass(frozen=True)
class Preview:
    columns: list[str]
    rows: list[dict]


def _read_csv_dict_rows(path: str, max_rows: int | None = None) -> tuple[list[str], list[dict]]:
    raw = Path(path).read_bytes()
    # Common encodings for Windows/broker exports
    for enc in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        # last resort: replace invalid bytes
        text = raw.decode("utf-8", errors="replace")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv.excel

    f = io.StringIO(text)
    reader = csv.DictReader(f, dialect=dialect)
    columns = [str(c) for c in (reader.fieldnames or [])]
    rows: list[dict] = []
    for i, row in enumerate(reader):
        if max_rows is not None and i >= max_rows:
            break
        rows.append({k: (v if v != "" else None) for k, v in row.items()})
    return columns, rows


def _safe_ext(filename: str) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    return ext if ext in {"csv", "xlsx", "xls"} else "bin"


def get_recon_dir(recon_id: str) -> str:
    root = current_app.config["UPLOAD_ROOT"]
    path = os.path.join(root, recon_id)
    os.makedirs(path, exist_ok=True)
    return path


def save_upload(recon_id: str, side: Side, storage_name: str, file_storage) -> str:
    recon_dir = get_recon_dir(recon_id)
    path = os.path.join(recon_dir, f"{side}-{storage_name}")
    file_storage.save(path)
    return path


def read_preview(path: str, max_rows: int = 50) -> Preview:
    ext = Path(path).suffix.lower()

    if ext == ".csv":
        columns, rows = _read_csv_dict_rows(path, max_rows=max_rows)
        return Preview(columns=columns, rows=rows)
    elif ext in {".xlsx", ".xls"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = next(it, None)
        if not headers:
            return Preview(columns=[], rows=[])
        columns = [str(h).strip() if h is not None else "" for h in headers]
        rows = []
        for i, row in enumerate(it):
            if i >= max_rows:
                break
            rows.append({columns[j]: (row[j] if j < len(row) else None) for j in range(len(columns))})
        return Preview(columns=columns, rows=rows)
    else:
        raise ValueError("Unsupported file type. Please upload CSV or Excel.")


