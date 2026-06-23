from __future__ import annotations

import os
import shutil
from datetime import datetime, timezone, timedelta
from pathlib import Path

from flask import Blueprint, request, send_file
from flask_jwt_extended import get_jwt, get_jwt_identity, jwt_required

from src.db import get_db
from src.mongo_ids import oid, oid_str
from src.routes.users import JURISDICTIONS
from src.utils.files import read_preview, save_upload
from src.utils.broker_templates import resolve_broker_template_key, run_build, run_stats, template_unavailable_message
from src.utils.recon_rows import (
    collect_normalized_rows,
    isin_summary_counts,
    split_rows_by_difference,
)
from src.utils.recon_cleanup import purge_reconciliation
from src.utils.break_comments import (
    active_break_ages_for_recon,
    break_comment_row_key,
    comment_for_export,
    isin_from_normalized_row,
    is_break_comment_row_key,
    is_latest_reviewed_for_account,
    sync_account_break_comments_on_build,
    upsert_account_break_comment,
)

from io import BytesIO
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

bp = Blueprint("reconciliations", __name__)


def _now():
    return datetime.now(timezone.utc)


def _to_username(value: str | None) -> str | None:
    if not value:
        return None
    s = str(value).strip()
    if not s:
        return None
    return s.split("@", 1)[0] if "@" in s else s


def _active_jurisdiction() -> str:
    j = (get_jwt() or {}).get("jurisdiction") or ""
    j = str(j).strip()
    return j if j in JURISDICTIONS else "ALL"


def _user_display_name(user: dict | None, fallback_identity: str) -> str:
    user = user or {}
    return (
        _to_username(user.get("fullName"))
        or _to_username(user.get("username"))
        or _to_username(user.get("email"))
        or _to_username(fallback_identity)
        or fallback_identity
    )


def _me(db) -> dict:
    return db["users"].find_one({"_id": oid(get_jwt_identity())}) or {}


def _is_operations(db) -> bool:
    """
    Operations permissions are driven by TEAM, not role.
    Backward compatible: treat legacy role=="operations" as operations.
    """
    me = _me(db)
    return str(me.get("team") or "") == "Operations" or str(me.get("role") or "") == "operations"


def _operations_eu_jurisdiction_filter() -> dict:
    """EU reconciliations; include ALL / legacy docs without jurisdiction."""
    return {
        "$or": [
            {"jurisdiction": "EU"},
            {"jurisdiction": "ALL"},
            {"jurisdiction": None},
            {"jurisdiction": {"$exists": False}},
        ]
    }


def _operations_reviewed_date_filter(days: int) -> dict:
    days = max(1, min(days, 7))
    today = _now().date()
    from_day = today if days == 1 else (today - timedelta(days=days - 1))
    start = datetime(from_day.year, from_day.month, from_day.day, tzinfo=timezone.utc)
    end = datetime(today.year, today.month, today.day, tzinfo=timezone.utc) + timedelta(days=1)
    return {
        "$or": [
            {"reviewedAt": {"$gte": start, "$lt": end}},
            {"reviewedAt": None, "updatedAt": {"$gte": start, "$lt": end}},
            {"reviewedAt": {"$exists": False}, "updatedAt": {"$gte": start, "$lt": end}},
        ]
    }


def _operations_reviewed_query(days: int = 1) -> dict:
    return {
        "status": "reviewed",
        "$and": [
            _operations_eu_jurisdiction_filter(),
            _operations_reviewed_date_filter(days),
        ],
    }


DASHBOARD_RECON_NOTE_PREFIX = "RECON|"


def _dashboard_recon_row_key(recon_id: str) -> str:
    return f"{DASHBOARD_RECON_NOTE_PREFIX}{recon_id}"


def _recon_display_name(
    recon: dict,
    *,
    broker_name: str | None = None,
    account_name: str | None = None,
) -> str:
    custom = (recon.get("name") or "").strip()
    if custom:
        return custom
    parts = [broker_name, account_name]
    left = " / ".join([p for p in parts if p])
    right = " • ".join([x for x in [recon.get("type"), recon.get("valueDate")] if x])
    label = " — ".join([x for x in [left, right] if x])
    return label or str(recon.get("type") or "Reconciliation")


def _safe_download_basename(label: str, fallback_id: str) -> str:
    import re

    safe = re.sub(r"[^\w\s-]", "", str(label or "")).strip()
    safe = re.sub(r"[\s_-]+", "_", safe)[:80]
    return safe or f"reconciliation_{fallback_id}"


def _jurisdiction_filter(active: str) -> dict:
    if active == "ALL":
        return {}
    return {"jurisdiction": {"$in": [active, "ALL", None]}}


def _dashboard_jurisdiction_filter(active: str) -> dict:
    """EU portal → EU only; US portal → US only; ALL portal → every jurisdiction."""
    if active == "ALL":
        return {}
    return {"jurisdiction": active}


def _today_utc_bounds() -> tuple[datetime, datetime]:
    today = _now().date()
    start = datetime(today.year, today.month, today.day, tzinfo=timezone.utc)
    end = start + timedelta(days=1)
    return start, end


def _performed_between_filter(start: datetime, end: datetime) -> dict:
    """Reconciliations built, submitted, or reviewed within [start, end)."""
    return {
        "$or": [
            {"submittedAt": {"$gte": start, "$lt": end}},
            {"completedAt": {"$gte": start, "$lt": end}},
            {"reviewedAt": {"$gte": start, "$lt": end}},
            {
                "status": "completed",
                "submittedAt": None,
                "updatedAt": {"$gte": start, "$lt": end},
            },
            {
                "status": "completed",
                "submittedAt": {"$exists": False},
                "updatedAt": {"$gte": start, "$lt": end},
            },
        ]
    }


def _dashboard_today_filter() -> dict:
    start, end = _today_utc_bounds()
    return _reviewed_between_filter(start, end)


def _reviewed_between_filter(start: datetime, end: datetime) -> dict:
    """Reconciliations reviewed within [start, end)."""
    return {
        "$or": [
            {"reviewedAt": {"$gte": start, "$lt": end}},
            {"reviewedAt": None, "updatedAt": {"$gte": start, "$lt": end}},
            {"reviewedAt": {"$exists": False}, "updatedAt": {"$gte": start, "$lt": end}},
        ]
    }


def _recon_matches_active_jurisdiction(recon: dict, active: str) -> bool:
    if active == "ALL":
        return True
    j = recon.get("jurisdiction")
    if j is None or j == "ALL":
        return True
    return str(j) == active


def _recon_matches_dashboard_jurisdiction(recon: dict, active: str) -> bool:
    if active == "ALL":
        return True
    return str(recon.get("jurisdiction") or "") == active


def _dashboard_workflow_statuses() -> list[str]:
    return ["reviewed"]


def _draft_statuses() -> list[str]:
    """In-progress reconciliations not yet submitted for review."""
    return ["draft", "uploaded", "completed"]


def _deletable_draft_statuses() -> set[str]:
    return set(_draft_statuses())


def _user_can_access_recon(db, recon: dict, *, for_dashboard_note: bool = False) -> bool:
    user_identity = get_jwt_identity()
    user_oid = oid(user_identity)
    active = _active_jurisdiction()

    if _is_operations(db):
        if recon.get("status") != "reviewed":
            return False
        return _recon_matches_dashboard_jurisdiction(recon, active)

    is_owner = recon.get("userId") == user_oid or str(recon.get("userId")) == str(user_identity)
    if is_owner:
        return True

    me = _me(db)
    if me.get("role") == "admin":
        return True

    if oid_str(recon.get("reviewerId")) == oid_str(user_oid):
        if recon.get("status") in {"submitted", "reviewed", "declined"}:
            return True

    if recon.get("status") == "reviewed" and _recon_matches_dashboard_jurisdiction(recon, active):
        return True

    if for_dashboard_note and recon.get("status") == "reviewed":
        return _recon_matches_dashboard_jurisdiction(recon, active)

    return False


def _recon_matches_view_jurisdiction(recon: dict, active: str) -> bool:
    if active == "ALL":
        return True
    j = recon.get("jurisdiction")
    return j in {active, "ALL", None}


def _get_recon_for_view(db, recon_id: str) -> dict | None:
    active = _active_jurisdiction()
    recon = db["reconciliations"].find_one({"_id": oid(recon_id)})
    if not recon:
        return None
    if not _recon_matches_view_jurisdiction(recon, active):
        return None
    if not _user_can_access_recon(db, recon):
        return None
    return recon


def _serialize_comment_history_item(item: dict) -> dict:
    out = {
        "break": item.get("break"),
        "comment": item.get("comment") or "",
        "createdAt": item.get("createdAt").isoformat() if item.get("createdAt") else None,
        "updatedBy": oid_str(item["updatedBy"]) if item.get("updatedBy") else None,
        "updatedByName": item.get("updatedByName"),
    }
    if item.get("difference") is not None:
        out["difference"] = item.get("difference")
    if item.get("archivedReason"):
        out["archivedReason"] = item.get("archivedReason")
    return out


def _resolve_username_from_updated_by(db, updated_by, users_by_id: dict | None = None) -> str | None:
    if not updated_by:
        return None
    try:
        # `updated_by` may be an ObjectId (mongo) or a string (legacy).
        from bson import ObjectId

        user_id = updated_by if isinstance(updated_by, ObjectId) else oid(str(updated_by))
        user = (users_by_id or {}).get(user_id)
        if user is None:
            user = db["users"].find_one({"_id": user_id})
    except Exception:
        user = None
    if not user:
        return None
    return _user_display_name(user, str(updated_by))


def _users_by_updated_by_ids(db, updated_by_ids: set) -> dict:
    if not updated_by_ids:
        return {}
    from bson import ObjectId

    ids: list = []
    for raw in updated_by_ids:
        try:
            ids.append(raw if isinstance(raw, ObjectId) else oid(str(raw)))
        except Exception:
            continue
    if not ids:
        return {}
    return {user["_id"]: user for user in db["users"].find({"_id": {"$in": ids}})}


def _lookup_maps_for_recons(db, docs: list[dict]) -> tuple[dict, dict, dict]:
    reviewer_ids: set = set()
    broker_ids: set = set()
    account_ids: set = set()
    for doc in docs:
        if doc.get("reviewerId"):
            reviewer_ids.add(doc["reviewerId"])
        if doc.get("brokerId"):
            broker_ids.add(doc["brokerId"])
        if doc.get("accountId"):
            account_ids.add(doc["accountId"])

    reviewers: dict = {}
    if reviewer_ids:
        for user in db["users"].find({"_id": {"$in": list(reviewer_ids)}}):
            reviewers[user["_id"]] = user
    brokers: dict = {}
    if broker_ids:
        for broker in db["brokers"].find({"_id": {"$in": list(broker_ids)}}):
            brokers[broker["_id"]] = broker
    accounts: dict = {}
    if account_ids:
        for account in db["accounts"].find({"_id": {"$in": list(account_ids)}}):
            accounts[account["_id"]] = account
    return reviewers, brokers, accounts


def _break_count_from_summary(doc: dict) -> int:
    summary = doc.get("summary") or {}
    return (
        int(summary.get("breaks") or 0)
        + int(summary.get("onlyOur") or 0)
        + int(summary.get("onlyCp") or 0)
    )


def _serialize_recon(
    doc: dict,
    db=None,
    *,
    reviewers: dict | None = None,
    brokers: dict | None = None,
    accounts: dict | None = None,
) -> dict:
    reviewer_name = None
    broker_name = None
    account_name = None
    if doc.get("reviewerId"):
        try:
            reviewer_user = (reviewers or {}).get(doc.get("reviewerId"))
            if reviewer_user is None and db is not None:
                reviewer_user = db["users"].find_one({"_id": doc.get("reviewerId")})
            reviewer_name = _user_display_name(reviewer_user, str(doc.get("reviewerId")))
        except Exception:
            reviewer_name = None
    if doc.get("brokerId") or doc.get("accountId"):
        try:
            b = (brokers or {}).get(doc.get("brokerId")) if doc.get("brokerId") else None
            a = (accounts or {}).get(doc.get("accountId")) if doc.get("accountId") else None
            if b is None and db is not None and doc.get("brokerId"):
                b = db["brokers"].find_one({"_id": doc.get("brokerId")})
            if a is None and db is not None and doc.get("accountId"):
                a = db["accounts"].find_one({"_id": doc.get("accountId")})
            broker_name = str((b or {}).get("name") or "") or None
            account_name = str((a or {}).get("name") or "") or None
        except Exception:
            broker_name = None
            account_name = None

    ops_comment_allowed = None
    if doc.get("status") == "reviewed" and db is not None:
        ops_comment_allowed = is_latest_reviewed_for_account(db, doc)

    return {
        "id": oid_str(doc["_id"]),
        "type": doc.get("type"),
        "status": doc.get("status"),
        "jurisdiction": doc.get("jurisdiction"),
        "valueDate": doc.get("valueDate"),
        "recDate": doc.get("recDate"),
        "performerName": _to_username(doc.get("performerName")),
        "reviewerId": oid_str(doc["reviewerId"]) if doc.get("reviewerId") else None,
        "reviewerName": _to_username(reviewer_name) if reviewer_name else None,
        "brokerId": oid_str(doc["brokerId"]) if doc.get("brokerId") else None,
        "accountId": oid_str(doc["accountId"]) if doc.get("accountId") else None,
        "brokerName": broker_name,
        "accountName": account_name,
        "declineReason": doc.get("declineReason"),
        "createdAt": doc.get("createdAt").isoformat() if doc.get("createdAt") else None,
        "updatedAt": doc.get("updatedAt").isoformat() if doc.get("updatedAt") else None,
        "reviewedAt": doc.get("reviewedAt").isoformat() if doc.get("reviewedAt") else None,
        "breakCount": _break_count_from_summary(doc),
        "ourFileName": (doc.get("ourFile") or {}).get("name"),
        "cpFileName": (doc.get("cpFile") or {}).get("name"),
        "name": (doc.get("name") or "").strip() or None,
        "dashboardComment": None,
        "opsCommentAllowed": ops_comment_allowed,
    }


@bp.get("")
@jwt_required()
def list_reconciliations():
    scope = (request.args.get("scope") or "").strip().lower()
    limit = int(
        request.args.get("limit")
        or (100 if scope in {"jurisdiction", "drafts", "today"} else 20)
    )
    limit = max(1, min(limit, 300 if scope in {"jurisdiction", "drafts", "today"} else 100))

    db = get_db()
    active = _active_jurisdiction()
    is_ops = _is_operations(db)

    if scope == "drafts":
        if is_ops:
            return {"items": []}
        q = {
            "userId": oid(get_jwt_identity()),
            "status": {"$in": _draft_statuses()},
        }
    elif scope == "today":
        q = {"recDate": _now().date().isoformat()}
        q.update(_dashboard_jurisdiction_filter(active))
    elif scope == "jurisdiction":
        q: dict = {"status": "reviewed"}
        q.update(_dashboard_jurisdiction_filter(active))
        days_window = request.args.get("days")
        if days_window:
            days_n = max(1, min(int(days_window), 7))
            today = _now().date()
            from_day = today - timedelta(days=days_n - 1)
            start = datetime(from_day.year, from_day.month, from_day.day, tzinfo=timezone.utc)
            end = datetime(today.year, today.month, today.day, tzinfo=timezone.utc) + timedelta(days=1)
            q["$and"] = [_reviewed_between_filter(start, end)]
        else:
            q["$and"] = [_dashboard_today_filter()]
    elif is_ops:
        days = int(request.args.get("days") or 1)
        q = _operations_reviewed_query(days)
    else:
        q = {"userId": oid(get_jwt_identity())}
        if active != "ALL":
            q["jurisdiction"] = {"$in": [active, "ALL", None]}

    # Do not let ?status= override fixed scopes (dashboard / drafts).
    if scope not in {"jurisdiction", "drafts", "today"}:
        status = (request.args.get("status") or "").strip()
        if status:
            parts = [x.strip() for x in status.split(",") if x.strip()]
            if parts:
                q["status"] = {"$in": parts}

    docs = list(
        db["reconciliations"]
        .find(q)
        .sort("createdAt", -1)
        .limit(limit)
    )
    reviewers, brokers, accounts = _lookup_maps_for_recons(db, docs)
    items = [
        _serialize_recon(d, reviewers=reviewers, brokers=brokers, accounts=accounts)
        for d in docs
    ]

    if scope == "jurisdiction" and docs:
        recon_ids = [d["_id"] for d in docs]
        row_keys = [_dashboard_recon_row_key(oid_str(rid)) for rid in recon_ids]
        notes_by_recon: dict[str, str] = {}
        for c in db["comments"].find(
            {"reconciliationId": {"$in": recon_ids}, "rowKey": {"$in": row_keys}}
        ):
            notes_by_recon[oid_str(c["reconciliationId"])] = c.get("comment") or ""
        for item in items:
            item["dashboardComment"] = notes_by_recon.get(item["id"]) or ""

    return {"items": items}


@bp.get("/review-queue")
@jwt_required()
def list_review_queue():
    limit = int(request.args.get("limit") or 50)
    limit = max(1, min(limit, 200))

    db = get_db()
    active = _active_jurisdiction()
    me = oid(get_jwt_identity())

    q: dict = {"reviewerId": me, "status": "submitted"}
    if active != "ALL":
        q["jurisdiction"] = {"$in": [active, "ALL", None]}

    docs = list(db["reconciliations"].find(q).sort("submittedAt", -1).limit(limit))
    reviewers, brokers, accounts = _lookup_maps_for_recons(db, docs)
    items = [
        _serialize_recon(d, reviewers=reviewers, brokers=brokers, accounts=accounts)
        for d in docs
    ]
    return {"items": items}


def _recon_search_haystack(doc: dict, item: dict) -> str:
    parts = [
        item.get("name"),
        item.get("type"),
        item.get("brokerName"),
        item.get("accountName"),
        item.get("status"),
        item.get("valueDate"),
        item.get("recDate"),
        item.get("performerName"),
        item.get("reviewerName"),
        item.get("jurisdiction"),
        _recon_display_name(
            doc,
            broker_name=item.get("brokerName"),
            account_name=item.get("accountName"),
        ),
        item.get("ourFileName"),
        item.get("cpFileName"),
    ]
    return " ".join(str(p) for p in parts if p).lower()


@bp.get("/search")
@jwt_required()
def search_reconciliations():
    q_text = (request.args.get("q") or "").strip()
    if len(q_text) < 2:
        return {"items": []}

    limit = max(1, min(int(request.args.get("limit") or 20), 50))
    db = get_db()
    active = _active_jurisdiction()
    is_ops = _is_operations(db)

    mongo_q: dict = {"status": "reviewed"}
    if is_ops:
        mongo_q.update(_dashboard_jurisdiction_filter(active))
    elif active != "ALL":
        mongo_q["jurisdiction"] = {"$in": [active, "ALL", None]}

    docs = list(
        db["reconciliations"]
        .find(mongo_q)
        .sort("reviewedAt", -1)
        .limit(400)
    )
    docs = [
        d
        for d in docs
        if _user_can_access_recon(db, d) and _recon_matches_view_jurisdiction(d, active)
    ]

    reviewers, brokers, accounts = _lookup_maps_for_recons(db, docs)
    needle = q_text.lower()
    matched: list[dict] = []
    for d in docs:
        item = _serialize_recon(d, reviewers=reviewers, brokers=brokers, accounts=accounts)
        if needle in _recon_search_haystack(d, item):
            matched.append(item)
            if len(matched) >= limit:
                break

    return {"items": matched}


@bp.get("/<recon_id>")
@jwt_required()
def get_reconciliation(recon_id: str):
    db = get_db()
    recon = _get_recon_for_view(db, recon_id)
    if not recon:
        return {"error": "Not found"}, 404
    return {"reconciliation": _serialize_recon(recon, db=db)}


@bp.delete("/<recon_id>")
@jwt_required()
def delete_reconciliation(recon_id: str):
    db = get_db()
    if _is_operations(db):
        return {"error": "Operations users cannot delete reconciliations"}, 403

    user_oid = oid(get_jwt_identity())
    recon_oid = oid(recon_id)
    recon = db["reconciliations"].find_one({"_id": recon_oid, "userId": user_oid})
    if not recon:
        return {"error": "Not found"}, 404

    if recon.get("status") not in _deletable_draft_statuses():
        return {"error": "Only draft reconciliations that have not been submitted can be deleted"}, 400

    purge_reconciliation(db, recon_oid)
    return {"ok": True}


@bp.post("")
@jwt_required()
def create_reconciliation():
    body = request.get_json(silent=True) or {}
    recon_type = (body.get("type") or "trade").strip().lower()
    if recon_type not in {"trade", "position", "fi"}:
        return {"error": "Invalid type"}, 400
    broker_id = (body.get("brokerId") or "").strip()
    account_id = (body.get("accountId") or "").strip()
    if not broker_id or not account_id:
        return {"error": "brokerId and accountId are required"}, 400
    value_date = (body.get("valueDate") or "").strip()
    if not value_date:
        return {"error": "valueDate is required"}, 400
    reviewer_id = (body.get("reviewerId") or "").strip()
    if not reviewer_id:
        return {"error": "reviewerId is required"}, 400

    db = get_db()
    if _is_operations(db):
        return {"error": "Operations users cannot create reconciliations"}, 403
    reviewer_user = db["users"].find_one({"_id": oid(reviewer_id)}) if reviewer_id else None
    if reviewer_user and str(reviewer_user.get("team") or "") == "Operations":
        return {"error": "Reviewer must be a Reconciliations team member"}, 400
    rec_date = _now().date().isoformat()
    performer_id = oid(get_jwt_identity())
    if oid_str(performer_id) == oid_str(oid(reviewer_id)):
        return {"error": "Performer and reviewer cannot be the same user"}, 400
    performer_user = db["users"].find_one({"_id": performer_id}) or {}
    performer_name = _user_display_name(performer_user, oid_str(performer_id))
    broker = db["brokers"].find_one({"_id": oid(broker_id)})
    if not broker:
        return {"error": "Broker not found"}, 404
    account = db["accounts"].find_one({"_id": oid(account_id), "brokerId": oid(broker_id)})
    if not account:
        return {"error": "Account not found for this broker. Create an account before starting a reconciliation."}, 404
    if not resolve_broker_template_key(broker, recon_type):
        return {"error": template_unavailable_message(broker, recon_type)}, 400
    active = _active_jurisdiction()
    doc = {
        "userId": oid(get_jwt_identity()),
        "jurisdiction": active,
        "type": recon_type,
        "brokerId": oid(broker_id),
        "accountId": oid(account_id),
        "valueDate": value_date,  # YYYY-MM-DD
        "recDate": rec_date,  # YYYY-MM-DD (server today)
        "performerId": performer_id,
        "performerName": performer_name,
        "reviewerId": oid(reviewer_id),
        "status": "draft",
        "createdAt": _now(),
        "updatedAt": _now(),
    }
    res = db["reconciliations"].insert_one(doc)
    return {"id": oid_str(res.inserted_id), "status": "draft", "type": recon_type}


@bp.post("/<recon_id>/upload")
@jwt_required()
def upload_files(recon_id: str):
    db = get_db()
    if _is_operations(db):
        return {"error": "Operations users cannot upload files"}, 403
    recon = db["reconciliations"].find_one({"_id": oid(recon_id), "userId": oid(get_jwt_identity())})
    if not recon:
        return {"error": "Not found"}, 404
    if recon.get("status") in {"submitted", "reviewed"}:
        return {"error": "Reconciliation is locked while under review"}, 400

    our_file = request.files.get("ourFile")
    cp_file = request.files.get("cpFile")
    if not our_file or not cp_file:
        return {"error": "Both ourFile and cpFile are required"}, 400

    our_name = Path(our_file.filename or "our.bin").name
    cp_name = Path(cp_file.filename or "cp.bin").name

    our_path = save_upload(recon_id, "our", our_name, our_file)
    cp_path = save_upload(recon_id, "cp", cp_name, cp_file)

    db["reconciliations"].update_one(
        {"_id": oid(recon_id)},
        {
            "$set": {
                "status": "uploaded",
                "ourFile": {"name": our_name, "path": our_path, "uploadedAt": _now()},
                "cpFile": {"name": cp_name, "path": cp_path, "uploadedAt": _now()},
                "updatedAt": _now(),
            }
        },
    )

    return {"ok": True}


@bp.get("/<recon_id>/preview")
@jwt_required()
def preview(recon_id: str):
    side = (request.args.get("side") or "").strip().lower()
    if side not in {"our", "cp"}:
        return {"error": "side must be our or cp"}, 400

    db = get_db()
    if _is_operations(db):
        return {"error": "Operations users cannot preview files"}, 403
    recon = db["reconciliations"].find_one({"_id": oid(recon_id), "userId": oid(get_jwt_identity())})
    if not recon:
        return {"error": "Not found"}, 404

    file_doc_key = "ourFile" if side == "our" else "cpFile"
    file_doc = recon.get(file_doc_key) or {}
    path = file_doc.get("path")
    if not path or not os.path.exists(path):
        return {"error": "File not uploaded"}, 400

    try:
        p = read_preview(path, max_rows=50)
    except Exception as e:
        return {"error": str(e)}, 400

    db["reconciliations"].update_one(
        {"_id": oid(recon_id)},
        {"$set": {f"{file_doc_key}.columns": p.columns, "updatedAt": _now()}},
    )

    return {"columns": p.columns, "rows": p.rows, "fileName": file_doc.get("name")}


@bp.get("/<recon_id>/stats")
@jwt_required()
def stats(recon_id: str):
    db = get_db()
    if _is_operations(db):
        return {"error": "Operations users cannot compute stats"}, 403
    recon = db["reconciliations"].find_one({"_id": oid(recon_id), "userId": oid(get_jwt_identity())})
    if not recon:
        return {"error": "Not found"}, 404

    our_path = (recon.get("ourFile") or {}).get("path")
    cp_path = (recon.get("cpFile") or {}).get("path")
    if not our_path or not cp_path:
        return {"error": "Files not uploaded"}, 400

    broker = db["brokers"].find_one({"_id": recon.get("brokerId")}) if recon.get("brokerId") else None
    recon_type = str(recon.get("type") or "trade")
    template_key = resolve_broker_template_key(broker, recon_type)
    if not template_key:
        return {"error": template_unavailable_message(broker, recon_type)}, 400

    try:
        s = run_stats(template_key, our_path, cp_path)
    except Exception as e:
        return {"error": str(e)}, 400

    return {"stats": s, "templateKey": template_key}


@bp.post("/<recon_id>/build")
@jwt_required()
def build(recon_id: str):
    db = get_db()
    if _is_operations(db):
        return {"error": "Operations users cannot build reconciliations"}, 403
    recon = db["reconciliations"].find_one({"_id": oid(recon_id), "userId": oid(get_jwt_identity())})
    if not recon:
        return {"error": "Not found"}, 404
    if recon.get("status") in {"submitted", "reviewed"}:
        return {"error": "Reconciliation is locked while under review"}, 400

    our_path = (recon.get("ourFile") or {}).get("path")
    cp_path = (recon.get("cpFile") or {}).get("path")
    if not our_path or not cp_path:
        return {"error": "Files not uploaded"}, 400

    broker = db["brokers"].find_one({"_id": recon.get("brokerId")}) if recon.get("brokerId") else None
    recon_type = str(recon.get("type") or "trade")
    template_key = resolve_broker_template_key(broker, recon_type)
    if not template_key:
        return {"error": template_unavailable_message(broker, recon_type)}, 400

    try:
        result = run_build(template_key, our_path, cp_path)
    except Exception as e:
        return {"error": str(e)}, 400

    db["reconciliations"].update_one(
        {"_id": oid(recon_id)},
        {
            "$set": {
                "status": "completed",
                "summary": result.summary,
                "completedAt": _now(),
                "updatedAt": _now(),
            }
        },
    )
    db["reconciliation_results"].update_one(
        {"reconciliationId": oid(recon_id)},
        {
            "$set": {
                "reconciliationId": oid(recon_id),
                "matched": result.matched,
                "breaks": result.breaks,
                "onlyOur": result.only_our,
                "onlyCp": result.only_cp,
                "summary": result.summary,
                "updatedAt": _now(),
            }
        },
        upsert=True,
    )

    updated_recon = db["reconciliations"].find_one({"_id": oid(recon_id)}) or recon
    _, break_rows = split_rows_by_difference(
        collect_normalized_rows(updated_recon, db["reconciliation_results"].find_one({"reconciliationId": oid(recon_id)}) or {})
    )
    sync_account_break_comments_on_build(db, updated_recon, break_rows)

    return {"ok": True, "summary": result.summary}


@bp.post("/<recon_id>/submit")
@jwt_required()
def submit_for_review(recon_id: str):
    db = get_db()
    if _is_operations(db):
        return {"error": "Operations users cannot submit reconciliations"}, 403
    recon = db["reconciliations"].find_one({"_id": oid(recon_id), "userId": oid(get_jwt_identity())})
    if not recon:
        return {"error": "Not found"}, 404
    if recon.get("status") not in {"completed", "declined"}:
        return {"error": "Reconciliation must be completed (or declined) before submitting for review"}, 400

    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    if len(name) > 120:
        return {"error": "Name must be 120 characters or fewer"}, 400

    update_fields: dict = {
        "status": "submitted",
        "submittedAt": _now(),
        "submittedBy": oid(get_jwt_identity()),
        "updatedAt": _now(),
    }
    if name:
        update_fields["name"] = name

    db["reconciliations"].update_one(
        {"_id": oid(recon_id)},
        {"$set": update_fields},
    )

    # Notify reviewer
    reviewer_id = recon.get("reviewerId")
    try:
        reviewer_user = db["users"].find_one({"_id": reviewer_id}) if reviewer_id else None
        reviewer_oid = reviewer_id if reviewer_id is not None else None
    except Exception:
        reviewer_user = None
        reviewer_oid = None

    if reviewer_oid is not None:
        performer_user = db["users"].find_one({"_id": oid(get_jwt_identity())}) or {}
        performer_name = _user_display_name(performer_user, str(get_jwt_identity()))
        reviewer_name = _user_display_name(reviewer_user, str(reviewer_oid)) if reviewer_user else None
        recon_label = name or _recon_display_name(recon)
        db["notifications"].insert_one(
            {
                "userId": reviewer_oid,
                "type": "reconciliation_assigned",
                "title": f"Review: {recon_label}",
                "body": f'{performer_name} submitted "{recon_label}" for your review.',
                "meta": {
                    "reconId": oid_str(oid(recon_id)),
                    "reconName": name or None,
                    "performerName": performer_name,
                    "reviewerName": reviewer_name,
                    "type": recon.get("type"),
                    "valueDate": recon.get("valueDate"),
                },
                "createdAt": _now(),
                "readAt": None,
            }
        )

    updated = db["reconciliations"].find_one({"_id": oid(recon_id)}) or {}
    return {"ok": True, "reconciliation": _serialize_recon(updated, db=db)}


@bp.patch("/<recon_id>/reviewer")
@jwt_required()
def update_reviewer(recon_id: str):
    db = get_db()
    if _is_operations(db):
        return {"error": "Operations users cannot change reviewers"}, 403

    recon = db["reconciliations"].find_one({"_id": oid(recon_id)})
    if not recon:
        return {"error": "Not found"}, 404

    user_oid = oid(get_jwt_identity())
    me = _me(db)
    is_admin = me.get("role") == "admin"
    is_owner = oid_str(recon.get("userId")) == oid_str(user_oid)
    if not is_admin and not is_owner:
        return {"error": "Only the performer can change the reviewer"}, 403

    if recon.get("status") not in {"submitted", "declined"}:
        return {"error": "Reviewer can only be changed while submitted or declined"}, 400

    body = request.get_json(silent=True) or {}
    reviewer_id = (body.get("reviewerId") or "").strip()
    if not reviewer_id:
        return {"error": "reviewerId is required"}, 400

    reviewer_user = db["users"].find_one({"_id": oid(reviewer_id)})
    if not reviewer_user:
        return {"error": "Reviewer not found"}, 404
    if str(reviewer_user.get("team") or "") == "Operations":
        return {"error": "Reviewer must be a Reconciliations team member"}, 400

    performer_id = recon.get("performerId") or recon.get("userId")
    if oid_str(performer_id) == oid_str(oid(reviewer_id)):
        return {"error": "Performer and reviewer cannot be the same user"}, 400

    new_reviewer_oid = oid(reviewer_id)
    if oid_str(recon.get("reviewerId")) == oid_str(new_reviewer_oid):
        return {"error": "Reviewer is already assigned"}, 400

    db["reconciliations"].update_one(
        {"_id": oid(recon_id)},
        {"$set": {"reviewerId": new_reviewer_oid, "updatedAt": _now()}},
    )

    if recon.get("status") == "submitted":
        performer_user = db["users"].find_one({"_id": performer_id}) or {}
        performer_name = _user_display_name(performer_user, str(performer_id))
        reviewer_name = _user_display_name(reviewer_user, reviewer_id)
        recon_label = _recon_display_name(recon)
        db["notifications"].insert_one(
            {
                "userId": new_reviewer_oid,
                "type": "reconciliation_assigned",
                "title": f"Review: {recon_label}",
                "body": f'{performer_name} assigned "{recon_label}" to you for review.',
                "meta": {
                    "reconId": oid_str(oid(recon_id)),
                    "reconName": recon.get("name") or None,
                    "performerName": performer_name,
                    "reviewerName": reviewer_name,
                    "type": recon.get("type"),
                    "valueDate": recon.get("valueDate"),
                },
                "createdAt": _now(),
                "readAt": None,
            }
        )

    updated = db["reconciliations"].find_one({"_id": oid(recon_id)}) or {}
    return {"ok": True, "reconciliation": _serialize_recon(updated, db=db)}


@bp.post("/<recon_id>/review")
@jwt_required()
def mark_reviewed(recon_id: str):
    db = get_db()
    user_oid = oid(get_jwt_identity())
    recon = db["reconciliations"].find_one({"_id": oid(recon_id)})
    if not recon:
        return {"error": "Not found"}, 404

    me = db["users"].find_one({"_id": user_oid}) or {}
    is_admin = me.get("role") == "admin"

    if not is_admin and oid_str(recon.get("reviewerId")) != oid_str(user_oid):
        return {"error": "Only the assigned reviewer can review this reconciliation"}, 403
    if recon.get("status") != "submitted":
        return {"error": "Reconciliation must be submitted before it can be reviewed"}, 400

    db["reconciliations"].update_one(
        {"_id": oid(recon_id)},
        {
            "$set": {
                "status": "reviewed",
                "reviewedAt": _now(),
                "reviewedBy": user_oid,
                "updatedAt": _now(),
            }
        },
    )
    updated = db["reconciliations"].find_one({"_id": oid(recon_id)}) or {}
    res = db["reconciliation_results"].find_one({"reconciliationId": oid(recon_id)})
    if res:
        _, break_rows = split_rows_by_difference(collect_normalized_rows(updated, res))
        sync_account_break_comments_on_build(db, updated, break_rows)
    return {"ok": True, "reconciliation": _serialize_recon(updated, db=db)}


@bp.post("/<recon_id>/decline")
@jwt_required()
def decline_review(recon_id: str):
    db = get_db()
    user_oid = oid(get_jwt_identity())
    recon = db["reconciliations"].find_one({"_id": oid(recon_id)})
    if not recon:
        return {"error": "Not found"}, 404

    me = db["users"].find_one({"_id": user_oid}) or {}
    is_admin = me.get("role") == "admin"
    if not is_admin and oid_str(recon.get("reviewerId")) != oid_str(user_oid):
        return {"error": "Only the assigned reviewer can decline this reconciliation"}, 403
    if recon.get("status") != "submitted":
        return {"error": "Reconciliation must be submitted before it can be declined"}, 400

    body = request.get_json(silent=True) or {}
    reason = str(body.get("reason") or "").strip()
    if not reason:
        return {"error": "reason is required"}, 400

    db["reconciliations"].update_one(
        {"_id": oid(recon_id)},
        {
            "$set": {
                "status": "declined",
                "declinedAt": _now(),
                "declinedBy": user_oid,
                "declineReason": reason,
                "updatedAt": _now(),
            }
        },
    )
    updated = db["reconciliations"].find_one({"_id": oid(recon_id)}) or {}
    return {"ok": True, "reconciliation": _serialize_recon(updated, db=db)}


@bp.post("/<recon_id>/redo")
@jwt_required()
def redo_reconciliation(recon_id: str):
    """
    Create a new draft reconciliation using the same metadata as a declined one.
    This keeps an audit trail (old declined rec remains unchanged).
    """
    db = get_db()
    if _is_operations(db):
        return {"error": "Operations users cannot redo reconciliations"}, 403
    user_identity = get_jwt_identity()
    user_oid = oid(user_identity)

    recon = db["reconciliations"].find_one({"_id": oid(recon_id)})
    if not recon:
        return {"error": "Not found"}, 404

    if recon.get("status") != "declined":
        return {"error": "Only declined reconciliations can be redone"}, 400

    is_owner = oid_str(recon.get("userId")) == oid_str(user_oid) or oid_str(recon.get("userId")) == str(user_identity)
    if not is_owner:
        return {"error": "You do not have access to this reconciliation"}, 403

    performer_user = db["users"].find_one({"_id": user_oid}) or {}
    performer_name = _user_display_name(performer_user, oid_str(user_oid))

    doc = {
        "userId": user_oid,
        "jurisdiction": recon.get("jurisdiction"),
        "type": recon.get("type"),
        "brokerId": recon.get("brokerId"),
        "accountId": recon.get("accountId"),
        "valueDate": recon.get("valueDate"),
        "recDate": _now().date().isoformat(),
        "performerId": user_oid,
        "performerName": performer_name,
        "reviewerId": recon.get("reviewerId"),
        "status": "draft",
        "createdAt": _now(),
        "updatedAt": _now(),
    }
    res = db["reconciliations"].insert_one(doc)
    return {"ok": True, "id": oid_str(res.inserted_id)}


def _normalize_rows_for_ui(recon: dict, res: dict) -> tuple[list[dict], list[dict], dict]:
    all_rows = collect_normalized_rows(recon, res)
    matched, breaks = split_rows_by_difference(all_rows)
    return matched, breaks, {}


@bp.get("/<recon_id>/export.xlsx")
@jwt_required()
def export_xlsx(recon_id: str):
    db = get_db()
    user_oid = oid(get_jwt_identity())

    recon = db["reconciliations"].find_one({"_id": oid(recon_id)})
    if not recon:
        return {"error": "Not found"}, 404

    me = db["users"].find_one({"_id": user_oid}) or {}
    is_admin = me.get("role") == "admin"
    is_owner = oid_str(recon.get("userId")) == oid_str(user_oid)
    is_reviewer = oid_str(recon.get("reviewerId")) == oid_str(user_oid)
    if not (is_admin or is_owner or is_reviewer):
        return {"error": "Permission denied"}, 403

    if recon.get("status") != "reviewed":
        return {"error": "Export is available only after the reconciliation is reviewed"}, 400

    res = db["reconciliation_results"].find_one({"reconciliationId": oid(recon_id)})
    if not res:
        return {"error": "No results yet"}, 400

    # Resolve names for checklist
    performer_display = _to_username(str(recon.get("performerName") or "")) or ""
    reviewer_user = db["users"].find_one({"_id": recon.get("reviewerId")}) if recon.get("reviewerId") else None
    reviewer_display = _user_display_name(reviewer_user, str(recon.get("reviewerId") or "")) if reviewer_user else ""

    broker_name = ""
    account_name = ""
    try:
        b = db["brokers"].find_one({"_id": recon.get("brokerId")}) if recon.get("brokerId") else None
        a = db["accounts"].find_one({"_id": recon.get("accountId")}) if recon.get("accountId") else None
        broker_name = str((b or {}).get("name") or "")
        account_name = str((a or {}).get("name") or "")
    except Exception:
        broker_name = ""
        account_name = ""

    matched_rows, break_rows, _ = _normalize_rows_for_ui(recon, res)

    comments_by_rowkey: dict[str, dict] = {}
    for c in db["comments"].find({"reconciliationId": oid(recon_id)}):
        comments_by_rowkey[c.get("rowKey")] = c

    break_ages = active_break_ages_for_recon(db, recon)

    def decorate(row: dict) -> dict:
        c = comment_for_export(comments_by_rowkey, row)
        br = c.get("break") or {}
        row = dict(row)
        row["Break Type"] = br.get("breakType") or ""
        row["Comments"] = (br.get("description") or c.get("comment") or "").strip()
        isin = isin_from_normalized_row(row)
        age_info = break_ages.get(break_comment_row_key(isin)) if isin else None
        row["Age (days)"] = age_info.get("breakAgeDays") if age_info else ""
        return row

    matched_rows = [decorate(r) for r in matched_rows]
    break_rows = [decorate(r) for r in break_rows]

    columns = [
        "Date",
        "Customer No",
        "Customer Name",
        "AT - ISIN",
        "AT Settled Quantity",
        "Broker ISIN",
        "Broker Settled Quantity",
        "Difference",
        "Age (days)",
        "Break Type",
        "Comments",
    ]

    wb = Workbook()
    ws_check = wb.active
    ws_check.title = "Checklist"
    ws_breaks = wb.create_sheet("Breaks")
    ws_matched = wb.create_sheet("Matched")

    # Professional Excel-like palette (light + muted accent)
    accent = "1F4E79"  # muted navy
    header_fill = PatternFill("solid", fgColor=accent)
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    row_alt_fill = PatternFill("solid", fgColor="F5F7FA")  # very light gray-blue
    row_base_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D0D7DE")  # soft gray borders
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_borders(ws):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_thin

    def is_number_col(col_name: str) -> bool:
        return col_name in {"AT Settled Quantity", "Broker Settled Quantity", "Difference"}

    def write_sheet(ws, rows: list[dict]):
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 110
        ws.append(columns)
        for col_idx in range(1, len(columns) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_align
            c.border = border_thin
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

        for r in rows:
            out = []
            for k in columns:
                v = r.get(k)
                if k == "Date" and isinstance(v, str) and v:
                    # store as date if possible
                    try:
                        y, m, d = v.split("-")
                        out.append(date(int(y), int(m), int(d)))
                    except Exception:
                        out.append(v)
                else:
                    out.append(v)
            ws.append(out)

        # Merge broker-side columns by Broker ISIN group (like UI rowspan)
        # Columns: F..J (Broker ISIN, Broker Qty, Difference, Break Type, Comments)
        def group_key(row_idx: int) -> str:
            v = ws.cell(row=row_idx, column=6).value  # Broker ISIN
            if v is None or str(v).strip() == "":
                v = ws.cell(row=row_idx, column=4).value  # AT - ISIN fallback
            return str(v or "").strip()

        merges: list[tuple[int, int]] = []
        start = 2
        while start <= ws.max_row:
            k = group_key(start)
            end = start
            while end + 1 <= ws.max_row and group_key(end + 1) == k:
                end += 1
            if k and end > start:
                merges.append((start, end))
            start = end + 1

        for s, e in merges:
            for col in range(6, 11):  # F..J
                try:
                    ws.merge_cells(start_row=s, start_column=col, end_row=e, end_column=col)
                except Exception:
                    pass

        # Formatting + borders
        for row_idx in range(2, ws.max_row + 1):
            fill = row_alt_fill if (row_idx % 2 == 0) else row_base_fill
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_name = columns[col_idx - 1]
                # alignment
                if col_name in {"AT - ISIN", "Broker ISIN", "Break Type"}:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif is_number_col(col_name):
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                # For merged broker-side columns, center-align like UI
                if col_idx in (6, 7, 8, 9, 10):
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                try:
                    cell.border = border_thin
                    cell.fill = fill
                    cell.font = Font(color="111827")  # near-black
                except Exception:
                    # merged cells may be read-only for some attrs in older openpyxl versions
                    pass

        # Ensure merged ranges have visible borders on all cells
        for s, e in merges:
            for col in range(6, 11):
                for r in range(s, e + 1):
                    try:
                        ws.cell(row=r, column=col).border = border_thin
                    except Exception:
                        pass

        # date format for Date col
        for cell in ws["A"][1:]:
            if isinstance(cell.value, date):
                cell.number_format = "mm/dd/yyyy"

        # numeric formats
        # E,G,H -> number
        for col_letter, fmt in [("E", "#,##0.#####"), ("G", "#,##0.#####"), ("H", "#,##0.###############")]:
            for cell in ws[col_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = fmt

        # simple width heuristics
        widths = {
            "A": 12,
            "B": 14,
            "C": 28,
            "D": 16,
            "E": 18,
            "F": 16,
            "G": 20,
            "H": 18,
            "I": 16,
            "J": 50,
        }
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

        # Make header row taller
        ws.row_dimensions[1].height = 26

    write_sheet(ws_breaks, break_rows)
    write_sheet(ws_matched, matched_rows)

    # Checklist sheet (first sheet)
    ws_check.sheet_view.showGridLines = False
    ws_check.sheet_view.zoomScale = 120
    ws_check.page_setup.fitToWidth = 1
    ws_check.page_setup.fitToHeight = 1
    ws_check.page_margins.left = 0.25
    ws_check.page_margins.right = 0.25
    ws_check.page_margins.top = 0.4
    ws_check.page_margins.bottom = 0.4

    # Wider, centered layout using A:E (with spacer column C)
    ws_check.column_dimensions["A"].width = 24
    ws_check.column_dimensions["B"].width = 34
    ws_check.column_dimensions["C"].width = 3
    ws_check.column_dimensions["D"].width = 24
    ws_check.column_dimensions["E"].width = 34

    # Title bar (accent)
    ws_check["A1"] = "Reconciliation Checklist"
    ws_check.merge_cells("A1:E1")
    ws_check["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_check["A1"].fill = PatternFill("solid", fgColor=accent)
    ws_check["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_check.row_dimensions[1].height = 34
    for col in range(1, 6):
        ws_check.cell(row=1, column=col).border = border_thin

    # Subheader (context line)
    recon_label = _recon_display_name(recon, broker_name=broker_name, account_name=account_name)
    ws_check["A2"] = f"{recon_label} • {str(recon.get('jurisdiction') or '')}"
    ws_check.merge_cells("A2:E2")
    ws_check["A2"].font = Font(bold=True, size=11, color="111827")
    ws_check["A2"].fill = PatternFill("solid", fgColor="E9EEF6")
    ws_check["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws_check.row_dimensions[2].height = 22
    for col in range(1, 6):
        ws_check.cell(row=2, column=col).border = border_thin

    section_fill = PatternFill("solid", fgColor="E9EEF6")
    section_font = Font(bold=True, size=11, color="111827")

    def section(title: str, row: int):
        ws_check[f"A{row}"] = title
        ws_check.merge_cells(f"A{row}:E{row}")
        ws_check[f"A{row}"].fill = section_fill
        ws_check[f"A{row}"].font = section_font
        ws_check[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws_check.row_dimensions[row].height = 20
        for col in range(1, 6):
            ws_check.cell(row=row, column=col).border = border_thin

    label_fill = PatternFill("solid", fgColor="F2F5FA")
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    label_font = Font(bold=True, color="111827")
    value_font = Font(color="111827")

    def put_pair(row: int, left_label: str, left_value, right_label: str | None = None, right_value=None):
        # left
        ws_check.cell(row=row, column=1, value=left_label)
        ws_check.cell(row=row, column=2, value=left_value)
        # spacer col 3
        ws_check.cell(row=row, column=3, value="")
        if right_label is not None:
            ws_check.cell(row=row, column=4, value=right_label)
            ws_check.cell(row=row, column=5, value=right_value)
        for col in range(1, 6):
            cell = ws_check.cell(row=row, column=col)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col in (1, 4):
                cell.fill = label_fill
                cell.font = label_font
            elif col in (2, 5):
                cell.fill = value_fill
                cell.font = value_font
            else:
                cell.fill = value_fill
        ws_check.row_dimensions[row].height = 20

    def iso_to_date(v):
        if isinstance(v, str) and v and "-" in v:
            try:
                y, m, d = v.split("-")
                return date(int(y), int(m), int(d))
            except Exception:
                return v
        return v

    # Run details
    section("Run details", 4)
    put_pair(5, "Reconciliation Type", recon.get("type") or "", "Status", recon.get("status") or "")
    put_pair(6, "Value Date", iso_to_date(recon.get("valueDate") or ""), "Rec Date", iso_to_date(recon.get("recDate") or ""))
    # date formatting
    for addr in ("B6", "E6"):
        if isinstance(ws_check[addr].value, date):
            ws_check[addr].number_format = "mm/dd/yyyy"
    put_pair(7, "Performer", performer_display, "Reviewer", reviewer_display)

    # Counts
    section("Counts", 9)
    breaks_cnt = int((res.get("summary") or {}).get("breaks") or 0)
    matched_cnt = int((res.get("summary") or {}).get("matched") or 0)
    total_isins = int((res.get("summary") or {}).get("totalIsins") or 0)
    put_pair(10, "No. of Breaks Reconciled", breaks_cnt, "No. of Matched", matched_cnt)
    put_pair(11, "Total ISINs", total_isins, "Decline reason", recon.get("declineReason") or "")

    ws_check.freeze_panes = "A5"

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"{_safe_download_basename(recon_label, recon_id)}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@bp.get("/<recon_id>/results")
@jwt_required()
def results(recon_id: str):
    db = get_db()
    recon = _get_recon_for_view(db, recon_id)
    if not recon:
        return {"error": "Not found"}, 404

    section = (request.args.get("section") or "all").strip().lower()
    include_matched = section in {"all", "matched"}
    include_breaks = section in {"all", "breaks"}

    res = db["reconciliation_results"].find_one({"reconciliationId": oid(recon_id)})
    if not res:
        return {"error": "No results yet. Build reconciliation first."}, 400

    # Matched = Difference is 0; Breaks = Difference has a value (incl. only-AT / only-CACEIS).
    all_rows = collect_normalized_rows(recon, res)
    matched, breaks = split_rows_by_difference(all_rows)
    isin_counts = isin_summary_counts(matched, breaks)

    # Load comments
    comment_docs = list(db["comments"].find({"reconciliationId": oid(recon_id)}))
    missing_updated_by: set = set()
    for c in comment_docs:
        history = c.get("history") or []
        if not history and (c.get("break") or c.get("comment")):
            history = [
                {
                    "break": c.get("break"),
                    "comment": c.get("comment") or "",
                    "createdAt": c.get("updatedAt"),
                    "updatedBy": c.get("updatedBy"),
                    "updatedByName": c.get("updatedByName"),
                }
            ]
        for h in history:
            if not h.get("updatedByName") and h.get("updatedBy"):
                missing_updated_by.add(h.get("updatedBy"))
    users_by_id = _users_by_updated_by_ids(db, missing_updated_by)

    comments = {}
    for c in comment_docs:
        history = c.get("history") or []
        if not history and (c.get("break") or c.get("comment")):
            history = [
                {
                    "break": c.get("break"),
                    "comment": c.get("comment") or "",
                    "createdAt": c.get("updatedAt"),
                    "updatedBy": c.get("updatedBy"),
                    "updatedByName": c.get("updatedByName"),
                }
            ]
        # Normalize history:
        # - backfill `updatedByName` for older rows
        # - carry forward the last known `break` so comment-only updates still show break context
        normalized_history = []
        last_break = None
        for h in history:
            if h.get("break"):
                last_break = h.get("break")
            elif last_break is not None:
                h = {**h, "break": last_break}
            if not h.get("updatedByName"):
                resolved = _resolve_username_from_updated_by(db, h.get("updatedBy"), users_by_id)
                if resolved:
                    h = {**h, "updatedByName": resolved}
            normalized_history.append(h)
        current_break = c.get("break") or last_break
        # Backward compatible:
        # - legacy: { comment: "..." }
        # - new: { break: { ...fields... }, comment: "..." }
        if current_break:
            comments[c["rowKey"]] = {
                "break": current_break,
                "comment": c.get("comment") or "",
                "history": [_serialize_comment_history_item(x) for x in normalized_history],
            }
        else:
            comments[c["rowKey"]] = {
                "comment": c.get("comment") or "",
                "history": [_serialize_comment_history_item(x) for x in normalized_history],
            }

    for row_key, age_info in active_break_ages_for_recon(db, recon).items():
        comments[row_key] = {**(comments.get(row_key) or {}), **age_info}

    payload: dict = {
        "summary": {
            **(res.get("summary") or {}),
            "matched": isin_counts["matched"],
            "breaks": isin_counts["breaks"],
        },
        "comments": comments,
    }
    if include_matched:
        payload["matched"] = matched
    if include_breaks:
        payload["breaks"] = breaks
    return payload


@bp.put("/<recon_id>/comments")
@jwt_required()
def upsert_comment(recon_id: str):
    db = get_db()
    user_identity = get_jwt_identity()
    recon_oid = oid(recon_id)
    user_oid = oid(user_identity)
    # Backward-compatible ownership check:
    # some old docs may store userId as string instead of ObjectId.
    recon = db["reconciliations"].find_one({"_id": recon_oid})
    if not recon:
        return {"error": "Not found"}, 404

    body = request.get_json(silent=True) or {}
    row_key = (body.get("rowKey") or "").strip()
    is_dashboard_note = row_key.startswith(DASHBOARD_RECON_NOTE_PREFIX)

    if not _user_can_access_recon(db, recon, for_dashboard_note=is_dashboard_note):
        return {"error": "You do not have access to this reconciliation"}, 403

    if (
        recon.get("status") in {"submitted", "reviewed"}
        and not _is_operations(db)
        and not is_dashboard_note
    ):
        return {"error": "Reconciliation is locked while under review"}, 400

    if _is_operations(db) and not is_dashboard_note:
        if recon.get("status") != "reviewed":
            return {"error": "Operations can only comment on reviewed reconciliations"}, 403
        if not is_latest_reviewed_for_account(db, recon):
            return {
                "error": "Operations can only comment on the latest reviewed reconciliation for this account",
            }, 403

    comment = body.get("comment") or ""
    break_payload = body.get("break")
    if not row_key:
        return {"error": "rowKey is required"}, 400

    # Validate break payload if present (used by Breaks comment modal)
    if break_payload is not None:
        if not isinstance(break_payload, dict):
            return {"error": "break must be an object"}, 400
        req_fields = ["breakType", "priority", "owner", "description"]
        for f in req_fields:
            v = (break_payload.get(f) or "").strip() if isinstance(break_payload.get(f), str) else break_payload.get(f)
            if not v:
                return {"error": f"{f} is required"}, 400

        # normalize optional fields
        if break_payload.get("mailSubject") is None:
            break_payload.pop("mailSubject", None)
        if break_payload.get("queryRaisedDate") is None:
            break_payload.pop("queryRaisedDate", None)
        # Keep legacy "comment" as description for quick display/search
        comment = break_payload.get("description") or comment

    history_item = {
        "break": break_payload,
        "comment": comment,
        "createdAt": _now(),
        "updatedBy": user_oid,
        "updatedByName": _user_display_name(db["users"].find_one({"_id": user_oid}), str(user_identity)),
    }

    db["comments"].update_one(
        {"reconciliationId": oid(recon_id), "rowKey": row_key},
        {
            "$set": {
                "reconciliationId": oid(recon_id),
                "rowKey": row_key,
                "comment": comment,
                "break": break_payload,
                "updatedAt": _now(),
                "updatedBy": user_oid,
                "updatedByName": _user_display_name(db["users"].find_one({"_id": user_oid}), str(user_identity)),
            },
            "$push": {"history": history_item},
        },
        upsert=True,
    )

    if is_break_comment_row_key(row_key):
        upsert_account_break_comment(
            db,
            recon,
            row_key,
            break_payload=break_payload,
            comment=comment,
            history_item=history_item,
        )

    return {"ok": True}

